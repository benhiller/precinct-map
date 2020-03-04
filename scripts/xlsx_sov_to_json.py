# Used for 2019 municipal election results

import sys
import json
from openpyxl import load_workbook
from collections import defaultdict

input_file = sys.argv[1]
output_file = sys.argv[2]

wb = load_workbook(input_file, read_only=True)

# { precinct: { registeredVoters, ballotsCast, contest: { candidate: votes }}}
precinct_data = defaultdict(lambda: defaultdict(dict))
contests = []

def parse_precincts(cell_value):
    raw_precincts = cell_value[4:]
    if raw_precincts.endswith('MB'):
        raw_precincts = raw_precincts[:-3]
    if '/' in raw_precincts:
        return raw_precincts.split('/')
    else:
        return [raw_precincts]

for idx, sheetname in enumerate(wb.sheetnames):
    ws = wb[sheetname]
    precincts = []
    # Turnout sheet
    if idx == 0:
        for row in ws.rows:
            if not row or not row[0] or not row[0].value:
                continue

            cell = row[0]
            if cell.value.startswith('PCT') or cell.value.startswith('Pct'):
                precincts = parse_precincts(cell.value)
            elif cell.value == 'Total':
                for precinct in precincts:
                    precinct_data[precinct]['registeredVoters'] = row[1].value
                    precinct_data[precinct]['ballotsCast'] = row[5].value
    # Other contest sheets
    else:
        contest = None
        col_to_candidate = {}
        for row_idx, row in enumerate(ws.rows):
            if row_idx == 1:
                contest = row[0].value.split('\n')[0].strip()
                contests.append(contest)
            elif row_idx == 3:
                seen_precinct = 0
                for cell in row:
                    if not cell.value:
                        continue
                    if cell.value == 'Precinct':
                        seen_precinct += 1
                        continue
                    if seen_precinct != 2:
                        continue
                    # TODO - ignoring write in candidates for now
                    if cell.value.strip() == 'Write-in' or cell.value.strip() == 'Total Votes':
                        break
                    else:
                        col_to_candidate[cell.column] = cell.value.split('\n')[0].strip()
            elif row_idx >= 6:
                if not row or not row[0] or not row[0].value:
                    continue

                cell = row[0]
                if cell.value.startswith('PCT') or cell.value.startswith('Pct'):
                    precincts = parse_precincts(cell.value)
                elif cell.value == 'Total':
                    for precinct in precincts:
                        for col in col_to_candidate:
                            precinct_data[precinct][contest][col_to_candidate[col]] = row[col - 1].value

f = open(output_file, 'w')
precinct_json = json.dumps({
    'contests': contests,
    'precinct_data': precinct_data,
})
f.write(precinct_json)
f.close()
